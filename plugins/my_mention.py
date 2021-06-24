from slackbot.bot import respond_to
from slackbot. bot import listen_to
from datetime import datetime
import openpyxl

    
@listen_to('上班')
def pounch_in(message):
    # 取得目前時間
    timestamp = datetime.now()
    date = timestamp.strftime('%Y/%m/%d')
    time = timestamp.strftime('%H:%M')
    print('上班時間登錄')

    # 回Slack
    message.send("上班囉 {}".format(time))

    # 儲時間到Excel
    wb = openpyxl.load_workbook("勤怠管理.xlsx")
    ws = wb.worksheets[0]
    max_row = ws.max_row

    # 日期
    ws[max_row + 1][0].value=date
    # 上班時間
    ws[max_row + 1][1].value=time

    # 存檔
    wb.save("勤怠管理.xlsx")

    print('上班時間登錄完成')


@listen_to('下班')
def pounch_out(message):
    # 取得目前時間
    timestamp = datetime.now()
    date = timestamp.strftime('%Y/%m/%d')
    time = timestamp.strftime('%H:%M')

    # 回Slack
    message.send("下班囉 {}".format(time))

        # 儲時間到Excel
    wb = openpyxl.load_workbook("勤怠管理.xlsx")
    ws = wb.worksheets[0]
    max_row = ws.max_row

    # 下班時間
    ws[max_row][2].value=time

    # 存檔
    wb.save("勤怠管理.xlsx")

    print('下班時間登錄完成')